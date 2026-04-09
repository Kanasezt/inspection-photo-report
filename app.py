from __future__ import annotations

import json
import os
import re
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageOps


# =========================
# CONFIG
# =========================
APP_TITLE = "Inspection Photo Report Tool"
TEMPLATE_FILENAME = "WIR Photo Tool V1.0.xlsx"
TARGET_SHEET = "TOOL"
MAX_UPLOADS = 20
RETENTION_HOURS = 24

# Excel image size based on user's screenshot
IMAGE_WIDTH_PX = int(round(2.49 * 96))   # ~239 px
IMAGE_HEIGHT_PX = int(round(1.87 * 96))  # ~180 px

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "app_data"
REPORT_DIR = DATA_DIR / "reports"
META_FILE = DATA_DIR / "report_index.json"
TMP_IMG_DIR = DATA_DIR / "tmp_images"
TEMPLATE_PATH = BASE_DIR / TEMPLATE_FILENAME

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "bmp", "webp"}

IMAGE_CELLS = [
    "B12", "G12", "G21", "B21", "B31", "G31",
    "B64", "G64", "G73", "B73", "B83", "G83",
    "B116", "G116", "G125", "B125", "B135", "G135",
    "B168", "G168", "G177", "B177", "B187", "G187",
    "B220", "G220", "G229", "B229", "B239", "G239",
    "B272", "G272", "G281", "B281", "B291", "G291",
]


# =========================
# STORAGE HELPERS
# =========================
def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    TMP_IMG_DIR.mkdir(parents=True, exist_ok=True)
    if not META_FILE.exists():
        META_FILE.write_text("[]", encoding="utf-8")


def load_index() -> List[Dict[str, Any]]:
    ensure_dirs()
    try:
        return json.loads(META_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_index(index_data: List[Dict[str, Any]]) -> None:
    META_FILE.write_text(json.dumps(index_data, indent=2, ensure_ascii=False), encoding="utf-8")


def now_local() -> datetime:
    return datetime.now()


def cleanup_expired_files() -> None:
    index_data = load_index()
    current_time = now_local()
    kept: List[Dict[str, Any]] = []

    for item in index_data:
        created_at = parse_iso(item.get("created_at"))
        expires_at = parse_iso(item.get("expires_at"))
        downloaded_at = parse_iso(item.get("downloaded_at")) if item.get("downloaded_at") else None
        path = REPORT_DIR / item.get("stored_name", "")

        remove_it = False
        if expires_at and current_time >= expires_at:
            remove_it = True
        elif downloaded_at and current_time >= downloaded_at + timedelta(minutes=10):
            remove_it = True
        elif not path.exists():
            remove_it = True

        if remove_it:
            try:
                if path.exists():
                    path.unlink()
            except Exception:
                pass
        else:
            kept.append(item)

    save_index(kept)

    # remove orphan temp images older than 24h
    for temp_file in TMP_IMG_DIR.glob("*"):
        try:
            modified = datetime.fromtimestamp(temp_file.stat().st_mtime)
            if current_time - modified > timedelta(hours=RETENTION_HOURS):
                temp_file.unlink(missing_ok=True)
        except Exception:
            pass


def parse_iso(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def register_report(display_name: str, stored_name: str, inspector_name: str, equipment_room: str) -> None:
    index_data = load_index()
    current_time = now_local()
    index_data.insert(0, {
        "id": str(uuid.uuid4()),
        "display_name": display_name,
        "stored_name": stored_name,
        "inspector_name": inspector_name,
        "equipment_room": equipment_room,
        "created_at": current_time.isoformat(timespec="seconds"),
        "expires_at": (current_time + timedelta(hours=RETENTION_HOURS)).isoformat(timespec="seconds"),
        "downloaded": False,
        "downloaded_at": None,
    })
    save_index(index_data)


def mark_report_downloaded(report_id: str) -> None:
    index_data = load_index()
    changed = False
    for item in index_data:
        if item.get("id") == report_id:
            item["downloaded"] = True
            item["downloaded_at"] = now_local().isoformat(timespec="seconds")
            changed = True
            break
    if changed:
        save_index(index_data)


def get_pending_reports() -> List[Dict[str, Any]]:
    cleanup_expired_files()
    pending = []
    for item in load_index():
        if item.get("downloaded"):
            continue
        path = REPORT_DIR / item.get("stored_name", "")
        if path.exists():
            pending.append(item)
    return pending


# =========================
# FILE / EXCEL HELPERS
# =========================
def sanitize_filename_component(text: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]", "", text or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "Inspector"


def build_output_filename(inspector_name: str) -> str:
    timestamp = now_local()
    date_part = timestamp.strftime("%Y-%m-%d")
    time_part = timestamp.strftime("%I%M")  # 4:58pm -> 0458
    safe_name = sanitize_filename_component(inspector_name)
    return f"Inspection_Report_{date_part}_{time_part} ({safe_name}).xlsx"


def process_uploaded_image(uploaded_file, seq_no: int) -> Path:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix.replace(".", "") not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {uploaded_file.name}")

    pil_image = Image.open(uploaded_file)
    pil_image = ImageOps.exif_transpose(pil_image)

    if pil_image.mode not in ("RGB", "RGBA"):
        pil_image = pil_image.convert("RGB")

    # Fit inside the target rectangle while keeping aspect ratio.
    if pil_image.mode == "RGBA":
        background = Image.new("RGB", pil_image.size, (255, 255, 255))
        background.paste(pil_image, mask=pil_image.split()[-1])
        pil_image = background
    else:
        pil_image = pil_image.convert("RGB")

    pil_image.thumbnail((IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX))

    temp_name = f"img_{uuid.uuid4().hex}_{seq_no:02d}.jpg"
    temp_path = TMP_IMG_DIR / temp_name
    pil_image.save(temp_path, format="JPEG", quality=88, optimize=True)
    return temp_path


def generate_report_bytes(inspector_name: str, equipment_room: str, uploaded_images) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template file not found: {TEMPLATE_FILENAME}")

    wb = load_workbook(TEMPLATE_PATH)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in template")

    ws = wb[TARGET_SHEET]
    ws["G8"] = equipment_room.strip()

    temp_paths: List[Path] = []

    try:
        for idx, uploaded in enumerate(uploaded_images[:len(IMAGE_CELLS)], start=1):
            temp_path = process_uploaded_image(uploaded, idx)
            temp_paths.append(temp_path)

            xl_img = XLImage(str(temp_path))
            xl_img.width = IMAGE_WIDTH_PX
            xl_img.height = IMAGE_HEIGHT_PX
            ws.add_image(xl_img, IMAGE_CELLS[idx - 1])

        output_stream = BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        return output_stream.getvalue()
    finally:
        for temp_path in temp_paths:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception:
                pass
        wb.close()


def save_report_to_disk(file_name: str, file_bytes: bytes) -> str:
    stored_name = f"{uuid.uuid4().hex}_{file_name}"
    out_path = REPORT_DIR / stored_name
    out_path.write_bytes(file_bytes)
    return stored_name


def format_file_size(num_bytes: int) -> str:
    kb = num_bytes / 1024
    if kb < 1024:
        return f"{kb:.1f} KB"
    return f"{kb / 1024:.2f} MB"


def reset_form() -> None:
    st.session_state["inspector_name"] = ""
    st.session_state["equipment_room"] = ""
    st.session_state["uploader_key"] = st.session_state.get("uploader_key", 0) + 1


# =========================
# UI
# =========================
st.set_page_config(page_title=APP_TITLE, page_icon="📷", layout="wide")
ensure_dirs()
cleanup_expired_files()

if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0
if "last_generated" not in st.session_state:
    st.session_state["last_generated"] = None
if "inspector_name" not in st.session_state:
    st.session_state["inspector_name"] = ""
if "equipment_room" not in st.session_state:
    st.session_state["equipment_room"] = ""

st.markdown(
    """
    <style>
    .main > div {padding-top: 2rem;}
    .block-container {max-width: 1200px;}
    .status-box {
        padding: 1rem 1.2rem;
        border-radius: 14px;
        background: linear-gradient(90deg, rgba(32,145,79,0.22), rgba(39,174,96,0.18));
        border: 1px solid rgba(80,200,120,0.25);
        margin-bottom: 1rem;
        font-size: 1.05rem;
    }
    .pending-card {
        border: 1px solid rgba(180,180,180,0.2);
        border-radius: 16px;
        padding: 0.9rem 1rem;
        margin-bottom: 0.7rem;
        background: rgba(255,255,255,0.02);
    }
    .small-muted {opacity: 0.75; font-size: 0.92rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📷 Inspection Photo Report Tool")
st.caption("Upload inspection photos and generate an Excel report from your template.")

if TEMPLATE_PATH.exists():
    st.markdown(f'<div class="status-box">Template loaded: <b>{TEMPLATE_FILENAME}</b></div>', unsafe_allow_html=True)
else:
    st.error(f"Template file not found: {TEMPLATE_FILENAME}. Put it in the same folder as app.py")

if st.session_state.get("last_generated"):
    st.success(f"Report generated: {st.session_state['last_generated']}")
    st.session_state["last_generated"] = None

with st.container(border=True):
    st.subheader("Generate new report")
    inspector_name = st.text_input(
        "Inspector Name",
        key="inspector_name",
        placeholder="Enter inspector name",
    )
    equipment_room = st.text_input(
        "Equipment / Room",
        key="equipment_room",
        placeholder="Enter equipment or room name",
    )

    uploaded_images = st.file_uploader(
        f"Select images (maximum {MAX_UPLOADS})",
        type=sorted(ALLOWED_EXTENSIONS),
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}",
        help="Supported: JPG, JPEG, PNG, BMP, WEBP",
    )

    image_count = len(uploaded_images) if uploaded_images else 0
    st.caption(f"Selected images: {image_count}/{MAX_UPLOADS}")

    generate_clicked = st.button("Generate Report", type="primary", use_container_width=False)

    if generate_clicked:
        if not TEMPLATE_PATH.exists():
            st.error("Template file is missing. Please upload or place the template file beside app.py first.")
        elif not inspector_name.strip():
            st.error("Please enter Inspector Name.")
        elif not equipment_room.strip():
            st.error("Please enter Equipment / Room.")
        elif not uploaded_images:
            st.error("Please upload at least 1 image.")
        elif len(uploaded_images) > MAX_UPLOADS:
            st.error(f"You can upload maximum {MAX_UPLOADS} images only.")
        else:
            try:
                with st.spinner("Generating Excel report..."):
                    file_name = build_output_filename(inspector_name.strip())
                    report_bytes = generate_report_bytes(
                        inspector_name=inspector_name.strip(),
                        equipment_room=equipment_room.strip(),
                        uploaded_images=uploaded_images,
                    )
                    stored_name = save_report_to_disk(file_name, report_bytes)
                    register_report(
                        display_name=file_name,
                        stored_name=stored_name,
                        inspector_name=inspector_name.strip(),
                        equipment_room=equipment_room.strip(),
                    )

                st.session_state["last_generated"] = file_name
                reset_form()
                st.rerun()
            except Exception as exc:
                st.error(f"Failed to generate report: {exc}")

st.markdown("---")
st.subheader("Pending files waiting for download")
pending_reports = get_pending_reports()

if not pending_reports:
    st.info("No pending report files. Files are kept for up to 24 hours.")
else:
    st.caption("Generated files that have not been downloaded yet. They will be cleared automatically after 24 hours.")
    for item in pending_reports:
        file_path = REPORT_DIR / item["stored_name"]
        created_at = parse_iso(item.get("created_at"))
        expires_at = parse_iso(item.get("expires_at"))
        size_text = format_file_size(file_path.stat().st_size) if file_path.exists() else "N/A"
        col1, col2 = st.columns([5, 2])
        with col1:
            st.markdown(
                f"""
                <div class="pending-card">
                    <div><b>{item['display_name']}</b></div>
                    <div class="small-muted">Inspector: {item.get('inspector_name', '-')} | Equipment/Room: {item.get('equipment_room', '-')}</div>
                    <div class="small-muted">Created: {created_at.strftime('%Y-%m-%d %I:%M %p') if created_at else '-'} | Expires: {expires_at.strftime('%Y-%m-%d %I:%M %p') if expires_at else '-'} | Size: {size_text}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with col2:
            with open(file_path, "rb") as f:
                downloaded = st.download_button(
                    label="Download Excel",
                    data=f.read(),
                    file_name=item["display_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{item['id']}",
                    use_container_width=True,
                )
            if downloaded:
                mark_report_downloaded(item["id"])
                st.rerun()

with st.expander("Current logic summary"):
    st.write("- Equipment / Room is written to sheet TOOL cell G8")
    st.write("- File name format: Inspection_Report_YYYY-MM-DD_HHMM (InspectorName).xlsx")
    st.write("- Time uses 12-hour format without colon, for example 4:58pm -> 0458")
    st.write(f"- Max upload: {MAX_UPLOADS} images")
    st.write("- Pending undownloaded files are shown below the form and auto-cleared after 24 hours")
